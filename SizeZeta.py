#!/usr/bin/env python3
import basic_tools
from collections import namedtuple

# 生成一个条形图，左侧为 Size，右侧为 Zeta
# 仅画图，不进行任何额外的处理

# See http://matplotlib.org/examples/api/barchart_demo.html
#     http://matplotlib.org/examples/api/two_scales.html

class SampleData(basic_tools.FrozenClass):
    def __init__(self):
        self.__init_slots__(["SampleName", "Size", "Zeta"])


def load_data(fp=basic_tools.HOME_PATH+"sample_data.xlsx") -> list:
    '''读取一个 xlsx 文件    '''
    import openpyxl

    from openpyxl import load_workbook
    wb = load_workbook(filename=fp, read_only=True)
    ws = wb[ wb.sheetnames[0] ]
    ret = []
    for r in tuple(ws.rows)[1:]:
        d = SampleData()
        d.SampleName = r[0].value
        d.Size       = r[1].value
        d.Zeta       = r[2].value
        ret.append(d)
    return ret

def draw_barchart(data):
    N = len(data)
    name = tuple(i.SampleName for i in data)
    size = tuple(i.Size for i in data)
    zeta = tuple(i.Zeta for i in data)

    import numpy as np
    import matplotlib.pyplot as plt

    fig, ax1 = plt.subplots()
    ind = np.arange(N)
    bar_width = 0.35

    rect1 = ax1.bar(ind-bar_width*0.5, size, bar_width,
            color='w', hatch='/', label="Size")
    ax1.set_ylabel("Size(nm)")
    ax1.set_xticks(ind)
    ax1.set_xticklabels(name)

    ax2 = ax1.twinx()
    #rect2 = ax2.bar(ind+bar_width, zeta,
            #bar_width, color='w', hatch='.', label="Zeta")
    line2 = ax2.plot(ind, zeta, color='black')
    ax2.set_ylabel("Zeta(mV)")
    ax2.axhline(0, color='black')

    import matplotlib.patches as mpatches
    import matplotlib.lines   as mlines
    ax1_patch = mpatches.Patch(facecolor='w', edgecolor='black', hatch='/', label="Size")
    ax2_line = mlines.Line2D([], [], color='black', label='Zeta')
    #plt.legend(handles=[ax1_patch, ax2_patch])
    plt.legend(handles=[ax1_patch, ax2_line])
    #plt.legend([rect1, line2])
    plt.show()

def main():
    data = load_data()
    draw_barchart(data)

if __name__ == '__main__':
    main()
else:
    raise Exception("SizeZeta.py shouldn't be used by other component")

